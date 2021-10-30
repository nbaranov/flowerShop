from abc import ABC, abstractmethod

import openpyxl


class Parser(ABC):
    @abstractmethod
    def parse(self, file):
        raise NotImplementedError()


class ExcelParser(Parser):
    def parse(self, file):
        work_book = openpyxl.load_workbook(file)
        sheet_names = work_book.sheetnames
        sheet = work_book[sheet_names[0]]

        data_table = []
        for row in range(2, sheet.max_row):
            data_row = [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column+1)]
            if data_row[0] is None:
                continue
            else:
                data_table.append(data_row)
        return data_table


class CsvParser(Parser):
    def parse(self, file):
        file_data = file.read().decode('utf-8')
        file_lines = file_data.split('\n')

        data_table = []
        for row in range(1, len(file_lines)):
            data_row = file_lines[row].split(',')
            if data_row[0] == '':  # проверяем не пустое ли поле External ID
                continue
            else:
                # костыль чтобы title, quantity и price обявить в None. пустая строка = ошибка при добавлении в db
                data_row = [None if item == '' else item for item in data_row]
                data_table.append(data_row)
        return data_table
